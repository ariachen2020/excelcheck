#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 格式比對工具 - Streamlit 版本
比對兩個 Excel 檔案的格式差異，包括：
- 欄位名稱和順序
- 資料類型
- 數值格式（小數位數、長度）
- 空值處理
- 儲存格格式
"""

import streamlit as st
import pandas as pd
import xlrd
from openpyxl import load_workbook
from pathlib import Path
from typing import Dict, Any, Tuple
import io
import re


class SumVerifier:
    """求和驗證器"""

    def __init__(self, excel_file):
        self.excel_file = excel_file

    def parse_cell_range(self, range_str: str) -> list:
        """
        解析儲存格範圍
        例如: "A1:A10" -> [A1, A2, ..., A10]
        """
        range_str = range_str.upper().strip()

        if ':' not in range_str:
            # 單個儲存格
            return [range_str]

        start_cell, end_cell = range_str.split(':')

        # 提取列和行
        start_col = re.match(r'[A-Z]+', start_cell).group()
        start_row = int(re.search(r'\d+', start_cell).group())

        end_col = re.match(r'[A-Z]+', end_cell).group()
        end_row = int(re.search(r'\d+', end_cell).group())

        # 轉換列字母為數字
        start_col_num = self._col_letter_to_num(start_col)
        end_col_num = self._col_letter_to_num(end_col)

        cells = []
        for col in range(start_col_num, end_col_num + 1):
            col_letter = self._col_num_to_letter(col)
            for row in range(start_row, end_row + 1):
                cells.append(f"{col_letter}{row}")

        return cells

    @staticmethod
    def _col_letter_to_num(col_letter: str) -> int:
        """將列字母轉換為數字 (A=1, B=2, ..., Z=26, AA=27)"""
        result = 0
        for char in col_letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result

    @staticmethod
    def _col_num_to_letter(col_num: int) -> str:
        """將列數字轉換為字母 (1=A, 2=B, ..., 26=Z, 27=AA)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result

    def verify_sum(self, cell_range: str, target_cell: str, sheet_index: int = 0) -> Dict[str, Any]:
        """
        驗證儲存格範圍的求和是否等於目標儲存格

        Args:
            cell_range: 儲存格範圍 (如 "A1:A10")
            target_cell: 目標儲存格 (如 "B1")
            sheet_index: 工作表索引 (默認為 0)

        Returns:
            驗證結果字典
        """
        try:
            # 嘗試用 openpyxl 讀取 (用於 .xlsx)
            try:
                wb = load_workbook(self.excel_file)
                ws = wb.worksheets[sheet_index]

                # 提取儲存格範圍中的值
                cells = self.parse_cell_range(cell_range)
                values = []
                for cell in cells:
                    try:
                        val = ws[cell].value
                        if val is not None and isinstance(val, (int, float)):
                            values.append(float(val))
                    except:
                        pass

                # 獲取目標儲存格的值
                target_val = ws[target_cell].value
                if target_val is None:
                    return {
                        'passed': False,
                        'error': f'目標儲存格 {target_cell} 為空',
                        'values': [],
                        'sum': 0,
                        'target': None
                    }

                target_val = float(target_val)
                sum_val = sum(values)

                # 檢查是否相等（允許浮點誤差）
                epsilon = 1e-9
                passed = abs(sum_val - target_val) < epsilon

                return {
                    'passed': passed,
                    'error': None,
                    'values': values,
                    'sum': sum_val,
                    'target': target_val,
                    'cell_range': cell_range,
                    'target_cell': target_cell,
                    'cells_count': len(values)
                }

            except Exception as e:
                # 如果 openpyxl 失敗，嘗試用 pandas
                try:
                    df = pd.read_excel(self.excel_file, sheet_name=sheet_index)
                    # 這種方法較難精確定位儲存格，返回錯誤
                    return {
                        'passed': False,
                        'error': '不支持此檔案格式，請使用 .xlsx 格式',
                        'values': [],
                        'sum': 0,
                        'target': None
                    }
                except:
                    return {
                        'passed': False,
                        'error': f'無法讀取檔案: {str(e)}',
                        'values': [],
                        'sum': 0,
                        'target': None
                    }

        except Exception as e:
            return {
                'passed': False,
                'error': f'驗證失敗: {str(e)}',
                'values': [],
                'sum': 0,
                'target': None
            }


class ExcelFormatChecker:
    """Excel 格式檢查器"""

    def __init__(self, correct_file, test_file):
        self.correct_file = correct_file
        self.test_file = test_file
        self.differences = []
        self.warnings = []

    def check_all(self) -> Dict[str, Any]:
        """執行所有檢查"""
        results = {
            'columns': self._check_columns(),
            'data_types': self._check_data_types(),
            'cell_formats': self._check_cell_formats(),
            'numeric_precision': self._check_numeric_precision(),
            'null_handling': self._check_null_handling(),
            'row_count': self._check_row_count(),
        }

        return results

    def _check_columns(self) -> Dict[str, Any]:
        """檢查欄位名稱和順序"""
        try:
            df_correct = pd.read_excel(self.correct_file, engine='xlrd')
            df_test = pd.read_excel(self.test_file, engine='xlrd')
        except Exception as e:
            return {
                'passed': False,
                'issues': [f"❌ 讀取檔案失敗: {str(e)}"],
                'correct_columns': [],
                'test_columns': []
            }

        correct_cols = list(df_correct.columns)
        test_cols = list(df_test.columns)

        issues = []

        # 檢查欄位數量
        if len(correct_cols) != len(test_cols):
            issues.append(f"❌ 欄位數量不同: 正確={len(correct_cols)}, 測試={len(test_cols)}")

        # 檢查欄位名稱
        missing_cols = set(correct_cols) - set(test_cols)
        extra_cols = set(test_cols) - set(correct_cols)

        if missing_cols:
            issues.append(f"❌ 缺少欄位: {missing_cols}")
        if extra_cols:
            issues.append(f"⚠️  多餘欄位: {extra_cols}")

        # 檢查欄位順序
        common_cols = [col for col in correct_cols if col in test_cols]
        for i, col in enumerate(common_cols):
            correct_idx = correct_cols.index(col)
            test_idx = test_cols.index(col)
            if correct_idx != test_idx:
                issues.append(f"⚠️  欄位 '{col}' 位置不同: 正確={correct_idx}, 測試={test_idx}")

        return {
            'passed': len(issues) == 0,
            'issues': issues,
            'correct_columns': correct_cols,
            'test_columns': test_cols
        }

    def _check_data_types(self) -> Dict[str, Any]:
        """檢查資料類型"""
        try:
            df_correct = pd.read_excel(self.correct_file, engine='xlrd')
            df_test = pd.read_excel(self.test_file, engine='xlrd')
        except Exception as e:
            return {
                'passed': False,
                'issues': [f"❌ 讀取檔案失敗: {str(e)}"]
            }

        issues = []
        common_cols = [col for col in df_correct.columns if col in df_test.columns]

        for col in common_cols:
            correct_dtype = df_correct[col].dtype
            test_dtype = df_test[col].dtype

            if correct_dtype != test_dtype:
                issues.append(f"❌ '{col}' 類型不同: 正確={correct_dtype}, 測試={test_dtype}")

        return {
            'passed': len(issues) == 0,
            'issues': issues
        }

    def _check_cell_formats(self) -> Dict[str, Any]:
        """檢查儲存格格式（使用 xlrd）"""
        try:
            wb_correct = xlrd.open_workbook(self.correct_file, formatting_info=True)
            wb_test = xlrd.open_workbook(self.test_file, formatting_info=True)
        except Exception as e:
            return {
                'passed': False,
                'issues': [f"❌ 讀取檔案失敗: {str(e)}"]
            }

        sheet_correct = wb_correct.sheet_by_index(0)
        sheet_test = wb_test.sheet_by_index(0)

        issues = []

        # 檢查第一筆資料（第2行，index=1）的格式
        if sheet_correct.nrows > 1 and sheet_test.nrows > 1:
            for col_idx in range(min(sheet_correct.ncols, sheet_test.ncols)):
                cell_correct = sheet_correct.cell(1, col_idx)
                cell_test = sheet_test.cell(1, col_idx)

                if cell_correct.ctype != cell_test.ctype:
                    col_name = sheet_correct.cell(0, col_idx).value
                    type_names = {0: "EMPTY", 1: "TEXT", 2: "NUMBER", 3: "DATE", 4: "BOOLEAN", 5: "ERROR"}
                    issues.append(
                        f"⚠️  欄 {col_idx} '{col_name}' 儲存格類型不同: "
                        f"正確={type_names.get(cell_correct.ctype, 'UNKNOWN')}, "
                        f"測試={type_names.get(cell_test.ctype, 'UNKNOWN')}"
                    )

        return {
            'passed': len(issues) == 0,
            'issues': issues
        }

    def _check_numeric_precision(self) -> Dict[str, Any]:
        """檢查數值精度和長度"""
        try:
            df_correct = pd.read_excel(self.correct_file, engine='xlrd')
            df_test = pd.read_excel(self.test_file, engine='xlrd')
        except Exception as e:
            return {
                'passed': False,
                'issues': [f"❌ 讀取檔案失敗: {str(e)}"]
            }

        issues = []
        common_cols = [col for col in df_correct.columns if col in df_test.columns]

        for col in common_cols:
            # 只檢查數值欄位
            if df_correct[col].dtype in ['int64', 'float64']:
                # 檢查第一筆非空值
                for idx in range(min(len(df_correct), len(df_test))):
                    val_correct = df_correct.iloc[idx][col]
                    val_test = df_test.iloc[idx][col]

                    if pd.notna(val_correct) and pd.notna(val_test):
                        # 檢查小數位數
                        str_correct = str(val_correct)
                        str_test = str(val_test)

                        if '.' in str_correct or '.' in str_test:
                            decimal_correct = len(str_correct.split('.')[-1]) if '.' in str_correct else 0
                            decimal_test = len(str_test.split('.')[-1]) if '.' in str_test else 0

                            if decimal_correct != decimal_test:
                                issues.append(
                                    f"⚠️  '{col}' 第 {idx+1} 筆小數位數不同: "
                                    f"正確={decimal_correct}位, 測試={decimal_test}位"
                                )

                        # 檢查總長度（去除小數點）
                        len_correct = len(str_correct.replace('.', ''))
                        len_test = len(str_test.replace('.', ''))

                        if len_correct != len_test:
                            issues.append(
                                f"⚠️  '{col}' 第 {idx+1} 筆長度不同: "
                                f"正確={len_correct}, 測試={len_test} "
                                f"(值: {val_correct} vs {val_test})"
                            )

                        # 檢查是否超過 15 位
                        if len_test > 15:
                            issues.append(
                                f"❌ '{col}' 第 {idx+1} 筆長度超過 15 位: {len_test} (值: {val_test})"
                            )

                        break  # 只檢查第一筆非空值

        return {
            'passed': len(issues) == 0,
            'issues': issues
        }

    def _check_null_handling(self) -> Dict[str, Any]:
        """檢查空值處理"""
        try:
            df_correct = pd.read_excel(self.correct_file, engine='xlrd')
            df_test = pd.read_excel(self.test_file, engine='xlrd')
        except Exception as e:
            return {
                'passed': False,
                'issues': [f"❌ 讀取檔案失敗: {str(e)}"]
            }

        issues = []
        common_cols = [col for col in df_correct.columns if col in df_test.columns]

        for col in common_cols:
            null_count_correct = df_correct[col].isna().sum()
            null_count_test = df_test[col].isna().sum()

            if null_count_correct != null_count_test:
                issues.append(
                    f"⚠️  '{col}' 空值數量不同: 正確={null_count_correct}, 測試={null_count_test}"
                )

        return {
            'passed': len(issues) == 0,
            'issues': issues
        }

    def _check_row_count(self) -> Dict[str, Any]:
        """檢查資料筆數"""
        try:
            df_correct = pd.read_excel(self.correct_file, engine='xlrd')
            df_test = pd.read_excel(self.test_file, engine='xlrd')
        except Exception as e:
            return {
                'passed': False,
                'issues': [f"❌ 讀取檔案失敗: {str(e)}"]
            }

        issues = []

        if len(df_correct) != len(df_test):
            issues.append(
                f"⚠️  資料筆數不同: 正確={len(df_correct)}, 測試={len(df_test)}"
            )

        return {
            'passed': len(issues) == 0,
            'issues': issues,
            'correct_rows': len(df_correct),
            'test_rows': len(df_test)
        }


def main():
    """主程式"""
    st.set_page_config(
        page_title="Excel 格式比對工具",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="expanded"
    )

    # 標題
    st.title("📊 Excel 格式比對工具")
    st.markdown("比對兩個 Excel 檔案的格式差異 | 驗證儲存格求和")

    # 使用 Tab 分頁
    tab1, tab2 = st.tabs(["格式比對", "求和驗證"])

    # 側邊欄說明
    with st.sidebar:
        st.header("使用說明")
        st.markdown("""
        ### Tab 1: 格式比對
        檢查兩個 Excel 檔案的格式差異
        - ✓ 欄位名稱和順序
        - ✓ 資料類型
        - ✓ 數值格式（小數位數、長度）
        - ✓ 空值處理
        - ✓ 儲存格格式
        - ✓ 資料筆數

        ### Tab 2: 求和驗證
        驗證儲存格範圍的求和是否等於目標儲存格
        - 支持 .xlsx 格式
        - 手動輸入儲存格位置
        - 實時計算和驗證
        """)

    # ===== Tab 1: 格式比對 =====
    with tab1:
        st.header("📋 格式比對")

        # 主要區域
        col1, col2 = st.columns(2)

        correct_file = None
        test_file = None

        with col1:
            st.subheader("正確的檔案")
            correct_uploaded = st.file_uploader(
                "上傳正確的 Excel 檔案",
                type=['xls', 'xlsx'],
                key="correct_file"
            )
            if correct_uploaded is not None:
                correct_file = correct_uploaded
                st.success(f"✓ 已上傳: {correct_uploaded.name}")

        with col2:
            st.subheader("待檢查的檔案")
            test_uploaded = st.file_uploader(
                "上傳待檢查的 Excel 檔案",
                type=['xls', 'xlsx'],
                key="test_file"
            )
            if test_uploaded is not None:
                test_file = test_uploaded
                st.success(f"✓ 已上傳: {test_uploaded.name}")

        # 檢查按鈕
        if correct_file and test_file:
            if st.button("🔍 開始比對", use_container_width=True):
                with st.spinner("正在比對檔案..."):
                    try:
                        checker = ExcelFormatChecker(correct_file, test_file)
                        results = checker.check_all()

                        # 顯示結果
                        st.markdown("---")
                        st.header("比對結果")

                        # 統計
                        total_checks = len(results)
                        passed_checks = sum(1 for r in results.values() if r['passed'])

                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("通過檢查", f"{passed_checks}/{total_checks}")
                        with col2:
                            st.metric("失敗檢查", f"{total_checks - passed_checks}/{total_checks}")
                        with col3:
                            if passed_checks == total_checks:
                                st.metric("狀態", "✅ 完全相同")
                            else:
                                st.metric("狀態", "❌ 有差異")

                        st.markdown("---")

                        # 詳細結果
                        check_names = {
                            'columns': '欄位名稱和順序',
                            'data_types': '資料類型',
                            'cell_formats': '儲存格格式',
                            'numeric_precision': '數值精度和長度',
                            'null_handling': '空值處理',
                            'row_count': '資料筆數'
                        }

                        for check_key, check_name in check_names.items():
                            result = results[check_key]

                            if result['passed']:
                                with st.expander(f"✅ {check_name}", expanded=False):
                                    st.success("檢查通過，無發現問題")
                                    if check_key == 'row_count' and 'correct_rows' in result:
                                        st.info(f"資料筆數: {result['correct_rows']}")
                            else:
                                with st.expander(f"❌ {check_name}", expanded=True):
                                    st.error("檢查失敗，發現以下問題:")
                                    for issue in result['issues']:
                                        st.write(issue)

                        st.markdown("---")

                        # 額外信息
                        if 'correct_columns' in results['columns'] and results['columns']['correct_columns']:
                            st.subheader("欄位對比")
                            col1, col2 = st.columns(2)
                            with col1:
                                st.write("**正確檔案的欄位:**")
                                for i, col in enumerate(results['columns']['correct_columns'], 1):
                                    st.write(f"{i}. {col}")
                            with col2:
                                st.write("**待檢查檔案的欄位:**")
                                for i, col in enumerate(results['columns']['test_columns'], 1):
                                    st.write(f"{i}. {col}")

                    except Exception as e:
                        st.error(f"發生錯誤: {str(e)}")
                        st.error("請確保上傳的是有效的 Excel 檔案")
        else:
            st.info("請上傳兩個 Excel 檔案開始比對")

    # ===== Tab 2: 求和驗證 =====
    with tab2:
        st.header("🧮 求和驗證")
        st.info("驗證儲存格範圍的求和是否等於指定的目標儲存格")

        # 上傳檔案
        uploaded_file = st.file_uploader(
            "上傳 Excel 檔案 (.xlsx)",
            type=['xlsx'],
            key="sum_verify_file"
        )

        if uploaded_file is not None:
            st.success(f"✓ 已上傳: {uploaded_file.name}")

            # 輸入欄位
            col1, col2 = st.columns(2)

            with col1:
                cell_range = st.text_input(
                    "儲存格範圍",
                    value="A1:A10",
                    help="例如: A1:A10, B2:B5 等"
                )

            with col2:
                target_cell = st.text_input(
                    "目標儲存格",
                    value="B1",
                    help="例如: B1, C10 等"
                )

            # 驗證按鈕
            if st.button("✓ 驗證求和", use_container_width=True):
                with st.spinner("正在驗證..."):
                    try:
                        verifier = SumVerifier(uploaded_file)
                        result = verifier.verify_sum(cell_range, target_cell)

                        # 顯示結果
                        st.markdown("---")
                        st.subheader("驗證結果")

                        if result['error']:
                            st.error(f"❌ 驗證失敗: {result['error']}")
                        else:
                            col1, col2, col3 = st.columns(3)

                            with col1:
                                st.metric("儲存格範圍", result['cell_range'])
                            with col2:
                                st.metric("計算的求和", f"{result['sum']:.2f}")
                            with col3:
                                st.metric("目標值", f"{result['target']:.2f}")

                            st.markdown("---")

                            if result['passed']:
                                st.success(
                                    f"✅ 驗證通過！\n\n"
                                    f"- 範圍: {result['cell_range']}\n"
                                    f"- 儲存格數量: {result['cells_count']}\n"
                                    f"- 求和結果: {result['sum']:.10g}\n"
                                    f"- 目標值 ({result['target_cell']}): {result['target']:.10g}\n"
                                    f"- 狀態: ✅ 相等"
                                )
                            else:
                                st.error(
                                    f"❌ 驗證失敗！\n\n"
                                    f"- 範圍: {result['cell_range']}\n"
                                    f"- 儲存格數量: {result['cells_count']}\n"
                                    f"- 求和結果: {result['sum']:.10g}\n"
                                    f"- 目標值 ({result['target_cell']}): {result['target']:.10g}\n"
                                    f"- 誤差: {abs(result['sum'] - result['target']):.10g}"
                                )

                            # 顯示詳細數值
                            with st.expander("查看詳細數值"):
                                st.write(f"範圍內的數值: {result['values']}")

                    except Exception as e:
                        st.error(f"發生錯誤: {str(e)}")
        else:
            st.info("請上傳 .xlsx 檔案開始驗證")


if __name__ == "__main__":
    main()
